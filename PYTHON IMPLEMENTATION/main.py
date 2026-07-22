# /// script
# requires-python = ">=3.10"
# dependencies = [
#     "cursor-sdk>=0.1.8",
#     "openpyxl>=3.1",
# ]
# ///
"""
main.py

Unattended website-grading automation, built on the official Cursor Python SDK
(package: cursor-sdk).

What it does
------------
For every row in the input spreadsheet where WEBSITE_GRADE is blank, this script:
  1. Sends a one-shot Cursor agent task: visit the company's website, and grade
     it using the "Company primary products" and "about Company who they are
     selling" columns as context.
  2. Parses the agent's structured answer (GOOD / MAYBE / UNABLE + a short reason).
  3. Writes the result back into WEBSITE_GRADE (and a short note into Comment).
  4. Saves progress after every batch, so the job is safe to stop and resume.

Runs multiple companies concurrently (each row is independent) using asyncio on
a single thread, with a per-row timeout and a small retry, so one slow or
broken site can't stall the batch. (Earlier versions used a thread pool, which
crashes on Windows with WinError 10038 when several threads share one local
agent connection -- asyncio avoids that by keeping everything on one thread.)

Requirements
------------
    Nothing to install by hand -- the "# /// script" block above lists the
    dependencies, and `uv run main.py` installs them automatically into an
    ephemeral environment the first time you run it.

Environment
-----------
    CURSOR_API_KEY must already be set in the environment. The SDK picks it up
    automatically -- nothing else to configure.

Usage
-----
    uv run main.py
    uv run main.py --limit 5                      (quick test on 5 rows first)
    uv run main.py --input "C:\\path\\companies.xlsx" --output "C:\\path\\out.xlsx"

    If --input / --output are omitted, the defaults below are used.
"""

from __future__ import annotations

import argparse
import asyncio
import json
import os
import re
import sys
import traceback
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from cursor_sdk import AsyncAgent, AsyncClient, AgentOptions, LocalAgentOptions

# ---------------------------------------------------------------------------
# Config (safe defaults -- override with CLI flags, see bottom of file)
# ---------------------------------------------------------------------------

# Default working folder and file names -- update here if the folder ever moves.
WORK_DIR = Path(r"D:\Upwork\Mac Clark\Alliance\Cursor Automation")
DEFAULT_INPUT = WORK_DIR / "companies.xlsx"
DEFAULT_OUTPUT = WORK_DIR / "companies_graded.xlsx"

MODEL_ID = "composer-2.5"          # Cursor's own agentic model -- fast + cheap for this
MAX_WORKERS = 6                    # how many companies to research at once
PER_ROW_TIMEOUT_SECONDS = 180      # give up on one company after this long
MAX_RETRIES = 1                    # one retry before falling back to UNABLE
SAVE_EVERY = 5                     # write progress to disk every N completed rows

COLUMN_NAMES = {
    "website": "Website Link",
    "products": "Company primary products",
    "about": "about Company who they are selling",
    "grade_out": "WEBSITE_GRADE",
    "comment_out": "Comment",
}

VALID_GRADES = {"GOOD", "MAYBE", "UNABLE"}


# ---------------------------------------------------------------------------
# Prompt construction
# ---------------------------------------------------------------------------

WHO_WE_ARE = """We are Alliance Test Equipment (alliancetesteq.com), based in Webster, MA.
We sell, rent, and lease used/refurbished electronic test & measurement
equipment -- oscilloscopes, spectrum/signal analyzers, signal generators,
meters, counters, power supplies, and wireless/optical test equipment, from
brands like Agilent/HP/Keysight, Tektronix, Anritsu, Fluke, and Rohde &
Schwarz. We also buy back customers' excess equipment and offer repair,
calibration, and consignment.

Our real customer is any company that has people on staff who need to
measure, test, calibrate, or diagnose electronic hardware -- not just any
company that happens to sell or use electronics."""


def build_prompt(website: str, products: str, about: str, examples: list[dict]) -> str:
    example_block = ""
    if examples:
        lines = []
        for ex in examples:
            lines.append(
                f'- Website: {ex["website"]}\n'
                f'  Products: {ex["products"]}\n'
                f'  About: {ex["about"]}\n'
                f'  -> Grade: {ex["grade"]}'
            )
        example_block = (
            "\nHere are a few examples of how these companies were graded before, "
            "for calibration:\n" + "\n".join(lines) + "\n"
        )

    return f"""You are screening a B2B lead for outreach on our behalf.

{WHO_WE_ARE}

Now research this prospect:

Website: {website}
Stated primary products: {products}
Who they say they sell to: {about}
{example_block}
Instructions:
- Actually open and read the website. Do not guess from the URL or company name alone.
- Do NOT write, edit, or create any files. Do NOT run shell commands. Only use your
  web browsing/fetch capability to view the site, then answer.
- The real question is NOT "is this company in electronics" -- it's "do they have
  people on staff who would actually use test & measurement equipment like an
  oscilloscope, signal generator, or power supply?" Look for signs of in-house
  engineering, R&D, manufacturing, calibration, diagnostics, or repair work on
  electronic hardware. A company that just resells or installs finished electronic
  products, with no visible engineering/test function, is NOT a good fit even if
  the word "electronics" is all over their site.
- Grade the company using ONLY one of these three values:
  - "GOOD"   -> clear evidence of in-house engineering, R&D, manufacturing,
                calibration, diagnostics, or repair work on electronic hardware --
                these are people who plausibly own or need test & measurement gear
  - "MAYBE"  -> electronics-adjacent, but unclear whether they do real in-house
                engineering/test work, or they mainly resell/install/integrate
                finished components without their own test function
  - "UNABLE" -> you could not load or meaningfully read the site (broken link,
                times out, blocked, parked domain, etc.)
- Keep the reason to one short sentence, and name the specific evidence you saw
  (e.g. "runs their own ADAS calibration lab" or "just resells finished units,
  no engineering team visible").

Respond with ONLY this exact JSON shape and nothing else -- no markdown, no code
fences, no extra commentary:
{{"grade": "GOOD", "reason": "one short sentence"}}
"""


# ---------------------------------------------------------------------------
# Result parsing
# ---------------------------------------------------------------------------

def parse_grade(raw_text: str) -> tuple[str, str]:
    """Best-effort extraction of {grade, reason} from the agent's final text."""
    text = raw_text.strip()

    # Try clean JSON first (agent was told not to use code fences, but be lenient).
    text_stripped = re.sub(r"^```(?:json)?|```$", "", text.strip(), flags=re.MULTILINE).strip()
    try:
        data = json.loads(text_stripped)
        grade = str(data.get("grade", "")).strip().upper()
        reason = str(data.get("reason", "")).strip()
        if grade in VALID_GRADES:
            return grade, reason or "(no reason given)"
    except (json.JSONDecodeError, AttributeError):
        pass

    # Fallback: look for a bare grade keyword anywhere in the text.
    match = re.search(r"\b(GOOD|MAYBE|UNABLE)\b", text.upper())
    if match:
        return match.group(1), text[:200]

    return "UNABLE", f"Could not parse agent response: {text[:200]}"


# ---------------------------------------------------------------------------
# One row's work
# ---------------------------------------------------------------------------

@dataclass
class RowTask:
    row_index: int          # 1-based openpyxl row number
    company: str
    website: str
    products: str
    about: str


async def grade_one_company(
    task: RowTask,
    examples: list[dict],
    client: AsyncClient,
    semaphore: asyncio.Semaphore,
) -> tuple[int, str, str]:
    """Runs a single one-shot Cursor agent call. Returns (row_index, grade, reason).

    All calls share one AsyncClient (one local bridge process, one event loop
    thread) -- this is what avoids the WinError 10038 socket crash that happens
    when multiple OS threads hit the same local bridge connection on Windows.
    The semaphore caps how many runs are in flight at once.
    """
    prompt = build_prompt(task.website, task.products, task.about, examples)

    async with semaphore:
        last_error = "unknown error"
        for attempt in range(MAX_RETRIES + 1):
            try:
                result = await asyncio.wait_for(
                    AsyncAgent.prompt(
                        prompt,
                        AgentOptions(
                            model=MODEL_ID,
                            api_key=os.environ.get("CURSOR_API_KEY"),
                            local=LocalAgentOptions(cwd=os.getcwd()),
                        ),
                        client=client,
                    ),
                    timeout=PER_ROW_TIMEOUT_SECONDS,
                )
                if result.status != "finished":
                    last_error = f"run status={result.status}"
                    continue
                grade, reason = parse_grade(result.result or "")
                return task.row_index, grade, reason
            except asyncio.TimeoutError:
                last_error = f"timed out after {PER_ROW_TIMEOUT_SECONDS}s"
            except Exception as exc:  # network errors, agent errors, etc.
                last_error = f"{type(exc).__name__}: {exc}"
                await asyncio.sleep(2)  # brief backoff before retrying

    return task.row_index, "UNABLE", f"Failed after {MAX_RETRIES + 1} attempt(s): {last_error}"


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------

async def run_batch(tasks: list[RowTask], examples: list[dict], args, wb, ws, col_grade, col_comment) -> int:
    # One local bridge process, one event loop, shared by every concurrent agent.
    client = await AsyncClient.launch_bridge(
        workspace=os.getcwd(),
        local=LocalAgentOptions(cwd=os.getcwd()),
        allow_api_key_env_fallback=True,
    )
    semaphore = asyncio.Semaphore(args.workers)
    completed = 0
    try:
        coros = {
            asyncio.ensure_future(grade_one_company(t, examples, client, semaphore)): t
            for t in tasks
        }
        for coro in asyncio.as_completed(list(coros.keys())):
            try:
                row_index, grade, reason = await coro
            except Exception as exc:
                # Should be rare -- grade_one_company already catches its own errors --
                # but fall back safely if something truly unexpected happens.
                row_index, grade, reason = -1, "UNABLE", f"Unexpected failure: {exc}"

            task = next((t for t in tasks if t.row_index == row_index), None)
            label = (task.company or task.website) if task else f"row {row_index}"

            if row_index != -1:
                ws.cell(row=row_index, column=col_grade, value=grade)
                if col_comment:
                    ws.cell(row=row_index, column=col_comment, value=reason)

            completed += 1
            print(f"[{completed}/{len(tasks)}] {label} -> {grade}: {reason}")

            if completed % SAVE_EVERY == 0:
                wb.save(args.output)
    finally:
        await client.aclose()

    return completed


def main() -> int:
    parser = argparse.ArgumentParser(description="Grade company websites with Cursor agents.")
    parser.add_argument("--input", default=str(DEFAULT_INPUT), help=f"Path to the input .xlsx (default: {DEFAULT_INPUT})")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT), help=f"Path to write the graded .xlsx (default: {DEFAULT_OUTPUT})")
    parser.add_argument("--sheet", default=None, help="Sheet name (default: active sheet)")
    parser.add_argument("--workers", type=int, default=MAX_WORKERS, help="Max companies researched at the same time")
    parser.add_argument("--limit", type=int, default=None, help="Only process the first N pending rows (for a test run)")
    args = parser.parse_args()

    if not os.environ.get("CURSOR_API_KEY"):
        print("ERROR: CURSOR_API_KEY is not set in the environment.", file=sys.stderr)
        return 1

    if not Path(args.input).exists():
        print(f"ERROR: input file not found: {args.input}", file=sys.stderr)
        return 1

    wb = load_workbook(args.input)
    ws = wb[args.sheet] if args.sheet else wb.active

    header = {cell.value: cell.column for cell in ws[1] if cell.value}
    missing = [name for name in [COLUMN_NAMES["website"], COLUMN_NAMES["products"],
                                  COLUMN_NAMES["about"], COLUMN_NAMES["grade_out"]]
               if name not in header]
    if missing:
        print(f"ERROR: missing expected column(s): {missing}", file=sys.stderr)
        return 1

    col_company = header.get("Company Name")
    col_website = header[COLUMN_NAMES["website"]]
    col_products = header[COLUMN_NAMES["products"]]
    col_about = header[COLUMN_NAMES["about"]]
    col_grade = header[COLUMN_NAMES["grade_out"]]
    col_comment = header.get(COLUMN_NAMES["comment_out"])

    # Pull a few already-graded rows as few-shot calibration examples.
    examples: list[dict] = []
    for row in ws.iter_rows(min_row=2):
        grade_val = row[col_grade - 1].value
        if grade_val and str(grade_val).strip().upper() in VALID_GRADES and len(examples) < 5:
            examples.append({
                "website": row[col_website - 1].value,
                "products": row[col_products - 1].value,
                "about": row[col_about - 1].value,
                "grade": str(grade_val).strip().upper(),
            })

    # Build the pending row list.
    tasks: list[RowTask] = []
    for row in ws.iter_rows(min_row=2):
        row_index = row[0].row
        grade_val = row[col_grade - 1].value
        if grade_val and str(grade_val).strip():
            continue  # already graded, skip
        website = row[col_website - 1].value
        if not website:
            continue  # nothing to grade
        tasks.append(RowTask(
            row_index=row_index,
            company=(row[col_company - 1].value if col_company else "") or "",
            website=str(website),
            products=str(row[col_products - 1].value or ""),
            about=str(row[col_about - 1].value or ""),
        ))

    if args.limit:
        tasks = tasks[: args.limit]

    print(f"{len(tasks)} companies to grade, up to {args.workers} at a time.")
    if not tasks:
        wb.save(args.output)
        print("Nothing to do. Saved a copy to", args.output)
        return 0

    completed = asyncio.run(run_batch(tasks, examples, args, wb, ws, col_grade, col_comment))

    wb.save(args.output)
    print(f"\nDone. {completed} rows graded. Saved to {args.output}")
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except KeyboardInterrupt:
        print("\nInterrupted -- progress already saved up to the last checkpoint.")
        sys.exit(130)
    except Exception:
        traceback.print_exc()
        sys.exit(1)