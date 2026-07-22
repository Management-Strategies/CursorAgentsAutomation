import openpyxl

path = r"c:\_netgit-three\check-target-net\tests\cursor_grading_web_tests\test_data\companies.xlsx"
wb = openpyxl.load_workbook(path)
ws = wb.active

headers = {c.value: c.column for c in ws[1] if c.value}
print("COLUMNS:", list(headers.keys()))

col_website = headers.get("Website Link")
col_about = headers.get("about Company")

for row in ws.iter_rows(min_row=2, max_row=8):
    website = row[col_website - 1].value if col_website else "?"
    print(f"Row {row[0].row}: website={website}")