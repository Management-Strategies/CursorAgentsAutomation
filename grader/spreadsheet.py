from __future__ import annotations

from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from grader.config import VALID_GRADES
from grader.models import ColumnMapping, ExampleRow, RowTask


def load_workbook_from_bytes(data: bytes) -> tuple[Workbook, Worksheet, str]:
    wb = load_workbook(BytesIO(data))
    ws = wb.active
    return wb, ws, ws.title


def get_headers(ws: Worksheet) -> list[str]:
    return [str(cell.value) for cell in ws[1] if cell.value is not None]


def header_map(ws: Worksheet) -> dict[str, int]:
    return {str(cell.value): cell.column for cell in ws[1] if cell.value is not None}


def ensure_output_columns(ws: Worksheet, mapping: ColumnMapping) -> dict[str, int]:
    headers = header_map(ws)
    if mapping.grade_out not in headers:
        col = ws.max_column + 1
        ws.cell(row=1, column=col, value=mapping.grade_out)
        headers[mapping.grade_out] = col
    if mapping.comment_out and mapping.comment_out not in headers:
        col = ws.max_column + 1
        ws.cell(row=1, column=col, value=mapping.comment_out)
        headers[mapping.comment_out] = col
    return headers


def worksheet_to_dataframe(ws: Worksheet) -> pd.DataFrame:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame()
    headers = [str(h) if h is not None else "" for h in rows[0]]
    data = rows[1:]
    return pd.DataFrame(data, columns=headers)


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _row_context(row: tuple, headers: dict[str, int], context_columns: list[str]) -> dict[str, str]:
    return {
        col: _cell_str(row[headers[col] - 1].value)
        for col in context_columns
        if col in headers
    }


def collect_examples(
    ws: Worksheet,
    mapping: ColumnMapping,
    max_examples: int = 5,
) -> list[ExampleRow]:
    headers = ensure_output_columns(ws, mapping)
    col_website = headers[mapping.website]
    col_grade = headers[mapping.grade_out]
    examples: list[ExampleRow] = []

    for row in ws.iter_rows(min_row=2):
        grade_val = row[col_grade - 1].value
        if not grade_val:
            continue
        grade = str(grade_val).strip().upper()
        if grade not in VALID_GRADES or len(examples) >= max_examples:
            continue
        website = row[col_website - 1].value
        if not website:
            continue
        examples.append(
            ExampleRow(
                website=str(website),
                context=_row_context(row, headers, mapping.context_columns),
                grade=grade,
            )
        )
    return examples


def build_pending_tasks(
    ws: Worksheet,
    mapping: ColumnMapping,
) -> list[RowTask]:
    headers = ensure_output_columns(ws, mapping)
    col_website = headers[mapping.website]
    col_grade = headers[mapping.grade_out]
    col_label = headers.get(mapping.label) if mapping.label else None

    tasks: list[RowTask] = []
    for row in ws.iter_rows(min_row=2):
        row_index = row[0].row
        grade_val = row[col_grade - 1].value
        if grade_val and _cell_str(grade_val):
            continue
        website = row[col_website - 1].value
        if not website:
            continue
        label = ""
        if col_label:
            label = _cell_str(row[col_label - 1].value)
        if not label:
            label = str(website)
        tasks.append(
            RowTask(
                task_id=f"row-{row_index}",
                row_index=row_index,
                label=label,
                website=str(website),
                context=_row_context(row, headers, mapping.context_columns),
            )
        )
    return tasks


def apply_result(
    ws: Worksheet,
    mapping: ColumnMapping,
    row_index: int,
    grade: str,
    reason: str,
) -> None:
    headers = ensure_output_columns(ws, mapping)
    col_grade = headers[mapping.grade_out]
    ws.cell(row=row_index, column=col_grade, value=grade)
    if mapping.comment_out:
        col_comment = headers[mapping.comment_out]
        ws.cell(row=row_index, column=col_comment, value=reason)


def workbook_to_bytes(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def guess_website_column(headers: list[str]) -> str | None:
    for header in headers:
        lower = header.lower()
        if "website" in lower or "url" in lower:
            return header
    return headers[0] if headers else None


def guess_context_columns(headers: list[str]) -> list[str]:
    legacy = [
        "Company primary products",
        "about Company who they are selling",
    ]
    found = [h for h in legacy if h in headers]
    if found:
        return found
    return [h for h in headers if h not in {"WEBSITE_GRADE", "Comment", "Company Name"}][:3]


def apply_dataframe_edits(
    ws: Worksheet,
    df: pd.DataFrame,
    mapping: ColumnMapping,
) -> None:
    headers = header_map(ws)
    col_grade = headers.get(mapping.grade_out)
    col_comment = headers.get(mapping.comment_out) if mapping.comment_out else None
    if not col_grade:
        return

    for idx, row in df.iterrows():
        row_index = idx + 2
        if mapping.grade_out in df.columns:
            grade = row.get(mapping.grade_out)
            if grade is not None and str(grade).strip():
                ws.cell(row=row_index, column=col_grade, value=str(grade).strip().upper())
        if col_comment and mapping.comment_out in df.columns:
            reason = row.get(mapping.comment_out)
            if reason is not None:
                ws.cell(row=row_index, column=col_comment, value=str(reason))
