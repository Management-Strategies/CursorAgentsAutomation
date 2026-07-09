# /// script
# requires-python = ">=3.10"
# dependencies = [
#     "cursor-sdk>=0.1.8",
#     "openpyxl>=3.1",
# ]
# ///
"""
main.py

Unattended website-grading automation, built on the official Cursor Python SDK
(package: cursor-sdk).

Usage
-----
    uv run main.py
    uv run main.py --limit 5
    uv run main.py --input "C:\\path\\companies.xlsx" --output "C:\\path\\out.xlsx"

Environment
-----------
    CURSOR_API_KEY must be set in the environment.
"""

from __future__ import annotations

import argparse
import asyncio
import os
import sys
import traceback
from pathlib import Path

from openpyxl import load_workbook

from grader.config import GradingConfig, LEGACY_COLUMN_NAMES
from grader.engine import run_batch_cli
from grader.models import ColumnMapping, RowTask
from grader.spreadsheet import build_pending_tasks, collect_examples, ensure_output_columns

WORK_DIR = Path(r"D:\Upwork\Mac Clark\Alliance\Cursor Automation")
DEFAULT_INPUT = WORK_DIR / "companies.xlsx"
DEFAULT_OUTPUT = WORK_DIR / "companies_graded.xlsx"
MAX_WORKERS = 6


def _legacy_mapping() -> ColumnMapping:
    return ColumnMapping(
        website=LEGACY_COLUMN_NAMES["website"],
        context_columns=[
            LEGACY_COLUMN_NAMES["products"],
            LEGACY_COLUMN_NAMES["about"],
        ],
        grade_out=LEGACY_COLUMN_NAMES["grade_out"],
        comment_out=LEGACY_COLUMN_NAMES["comment_out"],
        label=LEGACY_COLUMN_NAMES["label"],
    )


def _tasks_for_cli(ws, mapping: ColumnMapping) -> list[RowTask]:
    return build_pending_tasks(ws, mapping)


def main() -> int:
    parser = argparse.ArgumentParser(description="Grade company websites with Cursor agents.")
    parser.add_argument("--input", default=str(DEFAULT_INPUT), help=f"Input .xlsx (default: {DEFAULT_INPUT})")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT), help=f"Output .xlsx (default: {DEFAULT_OUTPUT})")
    parser.add_argument("--sheet", default=None, help="Sheet name (default: active sheet)")
    parser.add_argument("--workers", type=int, default=MAX_WORKERS, help="Max concurrent companies")
    parser.add_argument("--limit", type=int, default=None, help="Only process first N pending rows")
    args = parser.parse_args()

    if not os.environ.get("CURSOR_API_KEY"):
        print("ERROR: CURSOR_API_KEY is not set in the environment.", file=sys.stderr)
        return 1

    if not Path(args.input).exists():
        print(f"ERROR: input file not found: {args.input}", file=sys.stderr)
        return 1

    wb = load_workbook(args.input)
    ws = wb[args.sheet] if args.sheet else wb.active
    mapping = _legacy_mapping()
    ensure_output_columns(ws, mapping)

    config = GradingConfig.alliance_defaults()
    examples = collect_examples(ws, mapping, max_examples=config.max_examples)
    tasks = _tasks_for_cli(ws, mapping)

    if args.limit:
        tasks = tasks[: args.limit]

    print(f"{len(tasks)} companies to grade, up to {args.workers} at a time.")
    if not tasks:
        wb.save(args.output)
        print("Nothing to do. Saved a copy to", args.output)
        return 0

    def apply(row_index: int, grade: str, reason: str) -> None:
        headers = ensure_output_columns(ws, mapping)
        ws.cell(row=row_index, column=headers[mapping.grade_out], value=grade)
        if mapping.comment_out:
            ws.cell(row=row_index, column=headers[mapping.comment_out], value=reason)

    def checkpoint() -> None:
        wb.save(args.output)

    completed = asyncio.run(
        run_batch_cli(
            tasks,
            examples,
            config,
            args.workers,
            ws,
            apply,
            checkpoint_save=checkpoint,
        )
    )

    wb.save(args.output)
    print(f"\nDone. {completed} rows graded. Saved to {args.output}")
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except KeyboardInterrupt:
        print("\nInterrupted -- progress already saved up to the last checkpoint.")
        sys.exit(130)
    except Exception:
        traceback.print_exc()
        sys.exit(1)
