"""
streamlit_app.py

Streamlit front end for the website-grading automation. Wraps grading_engine.py
(the background streaming batch engine) and prompts.py (prompt construction)
around four screens: upload, configure, live run ("Mission Control"), and
review/export. See the plan doc for the full design rationale.
"""

from __future__ import annotations

import os
import time
from collections import deque
from io import BytesIO
from queue import Empty

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from grading_engine import EngineConfig, GradingEngine, RowTask
from prompts import DEFAULT_GRADE_RUBRIC, DEFAULT_WHO_WE_ARE, VALID_GRADES

REQUIRED_COLUMNS = ["Website Link", "WEBSITE_GRADE"]
OUTPUT_COLUMNS = ["WEBSITE_GRADE", "Comment"]
MAX_EXAMPLES = 5
GRADE_BADGE_COLOR = {"GOOD": "green", "MAYBE": "orange", "UNABLE": "gray"}
PHASE_PREFIX = {"tool-call-started": "Fetching", "thinking-delta": "Reasoning", "text-delta": "Drafting"}
MERGEABLE_DELTA_KINDS = {"thinking-delta", "text-delta"}  # streamed text, not discrete actions like tool calls

st.set_page_config(page_title="Lead Website Grader", page_icon="🧭", layout="wide")


# ---------------------------------------------------------------------------
# 0. Access gate
# ---------------------------------------------------------------------------

def safe_secret(key: str) -> str | None:
    """st.secrets raises StreamlitSecretNotFoundError if no secrets.toml
    exists anywhere at all, rather than just behaving like an empty dict."""
    try:
        return st.secrets.get(key)
    except Exception:
        return None


def check_password() -> bool:
    configured = safe_secret("APP_PASSWORD")
    if not configured:
        return True  # no password configured -- dev mode, gate is a no-op
    if st.session_state.get("authed"):
        return True
    st.title("🧭 Lead Website Grader")
    pw = st.text_input("Password", type="password")
    if st.button("Enter") or pw:
        if pw == configured:
            st.session_state["authed"] = True
            st.rerun()
        elif pw:
            st.error("Wrong password.")
    return False


def get_api_key() -> str | None:
    return safe_secret("CURSOR_API_KEY") or os.environ.get("CURSOR_API_KEY")


# ---------------------------------------------------------------------------
# Helpers shared across screens
# ---------------------------------------------------------------------------

def load_sheet(file) -> None:
    wb = load_workbook(file)
    ws = wb.active
    header = {cell.value: cell.column for cell in ws[1] if cell.value}

    missing = [c for c in REQUIRED_COLUMNS if c not in header]
    if missing:
        st.error(f"Missing required column(s): {missing}. Please check the file and re-upload.")
        return
    if "Comment" not in header:
        # Add it so grading has somewhere to write a reason.
        col = ws.max_column + 1
        ws.cell(row=1, column=col, value="Comment")
        header["Comment"] = col

    st.session_state.wb = wb
    st.session_state.ws = ws
    st.session_state.header = header
    st.session_state.all_columns = list(header.keys())
    st.session_state.col_company = header.get("Company Name")
    st.session_state.col_website = header["Website Link"]
    st.session_state.col_grade = header["WEBSITE_GRADE"]
    st.session_state.col_comment = header["Comment"]

    defaults = [c for c in ("Website Link", "Company primary products", "about Company who they are selling") if c in header]
    st.session_state.input_columns = defaults
    st.session_state.who_we_are = DEFAULT_WHO_WE_ARE
    st.session_state.grade_rubric = DEFAULT_GRADE_RUBRIC
    st.session_state.stage = "setup"


def dataframe_from_ws() -> pd.DataFrame:
    ws = st.session_state.ws
    columns = st.session_state.all_columns
    header = st.session_state.header
    rows = []
    for row in ws.iter_rows(min_row=2):
        if not row[header["Website Link"] - 1].value:
            continue
        rows.append({col: row[header[col] - 1].value for col in columns})
    return pd.DataFrame(rows)


def build_examples() -> list[dict]:
    ws = st.session_state.ws
    header = st.session_state.header
    col_grade = st.session_state.col_grade
    col_website = st.session_state.col_website
    input_columns = st.session_state.input_columns

    examples = []
    for row in ws.iter_rows(min_row=2):
        grade_val = row[col_grade - 1].value
        if grade_val and str(grade_val).strip().upper() in VALID_GRADES and len(examples) < MAX_EXAMPLES:
            examples.append({
                "website": row[col_website - 1].value,
                "fields": {col: row[header[col] - 1].value for col in input_columns},
                "grade": str(grade_val).strip().upper(),
            })
    return examples


def build_tasks(only_ungraded: bool, row_limit: int | None, only_row_indices: set[int] | None = None) -> list[RowTask]:
    ws = st.session_state.ws
    header = st.session_state.header
    col_company = st.session_state.col_company
    col_website = st.session_state.col_website
    col_grade = st.session_state.col_grade
    input_columns = st.session_state.input_columns

    tasks = []
    for row in ws.iter_rows(min_row=2):
        row_index = row[0].row
        if only_row_indices is not None and row_index not in only_row_indices:
            continue
        website = row[col_website - 1].value
        if not website:
            continue
        grade_val = row[col_grade - 1].value
        if only_row_indices is None and only_ungraded and grade_val and str(grade_val).strip():
            continue
        tasks.append(RowTask(
            row_index=row_index,
            company=(row[col_company - 1].value if col_company else "") or str(website),
            website=str(website),
            fields={col: str(row[header[col] - 1].value or "") for col in input_columns},
        ))
    if row_limit:
        tasks = tasks[:row_limit]
    return tasks


def start_run(tasks: list[RowTask]) -> None:
    examples = build_examples()
    config = EngineConfig(
        workers=st.session_state.workers,
        who_we_are=st.session_state.who_we_are,
        grade_rubric=st.session_state.grade_rubric,
        api_key=get_api_key(),
        cwd=os.getcwd(),
    )
    engine = GradingEngine(tasks, examples, config)
    st.session_state.engine = engine
    st.session_state.worker_slots = {i: None for i in range(config.workers)}
    st.session_state.activity_feed = deque(maxlen=len(tasks) or 1)  # keep every result for this run
    st.session_state.run_total = len(tasks)
    st.session_state.run_start_time = time.time()
    st.session_state.run_counts = {"GOOD": 0, "MAYBE": 0, "UNABLE": 0}
    st.session_state.run_done = 0
    st.session_state.run_finished = False
    engine.start()
    st.session_state.stage = "running"


def write_result_to_sheet(row_index: int, grade: str, reason: str) -> None:
    ws = st.session_state.ws
    ws.cell(row=row_index, column=st.session_state.col_grade, value=grade)
    ws.cell(row=row_index, column=st.session_state.col_comment, value=reason)


def workbook_bytes() -> bytes:
    buf = BytesIO()
    st.session_state.wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# 1 & 2. Upload + configure
# ---------------------------------------------------------------------------

def render_upload() -> None:
    st.title("🧭 Lead Website Grader")
    st.markdown(
        "This app grades a spreadsheet of leads for you: for every company, a Cursor AI "
        "agent **actually visits their website**, reads it, and decides whether they're a "
        "good fit -- **GOOD**, **MAYBE**, or **UNABLE** -- with a one-line reason, written "
        "straight back into your sheet."
    )
    st.markdown(
        "**How it works:**\n"
        "1. **Upload** your `.xlsx` below -- you'll see the entire sheet, unchanged, so you "
        "can check it's the right file.\n"
        "2. **Configure** in the sidebar: pick exactly which columns the AI reads as context "
        "(the website itself is always visited), and edit the grading criteria -- who you are "
        "and what counts as GOOD/MAYBE/UNABLE -- in plain English.\n"
        "3. **Run** it and watch each agent work in real time -- what site it's reading, what "
        "it's reasoning about, and what it decides.\n"
        "4. **Review** the results, override any grade by hand, and download the finished "
        "spreadsheet."
    )
    uploaded = st.file_uploader("Upload companies .xlsx", type=["xlsx"])
    if uploaded is not None and st.session_state.get("uploaded_name") != uploaded.name:
        st.session_state.uploaded_name = uploaded.name
        load_sheet(uploaded)
        st.rerun()


def render_setup() -> None:
    st.title("🧭 Lead Website Grader")
    df = dataframe_from_ws()
    graded = df["WEBSITE_GRADE"].apply(lambda v: bool(str(v).strip()) if v is not None else False).sum()
    c1, c2, c3 = st.columns(3)
    c1.metric("Total rows", len(df))
    c2.metric("Already graded", int(graded))
    c3.metric("Pending", len(df) - int(graded))

    st.subheader("Full sheet")
    st.dataframe(df, use_container_width=True, height=700)

    with st.sidebar:
        st.header("Configure")
        candidate_columns = [c for c in st.session_state.all_columns if c not in OUTPUT_COLUMNS]
        st.session_state.input_columns = st.multiselect(
            "What should the AI read?", options=candidate_columns,
            default=[c for c in st.session_state.input_columns if c in candidate_columns],
            help="The site behind Website Link is always visited regardless of this selection -- "
                 "these are the columns given to the AI as extra text context alongside it.",
        )

        with st.expander("Grading criteria", expanded=False):
            st.session_state.who_we_are = st.text_area("Who we are", value=st.session_state.who_we_are, height=180)
            st.session_state.grade_rubric = st.text_area("What GOOD / MAYBE / UNABLE mean", value=st.session_state.grade_rubric, height=220)
            if st.button("Reset to defaults"):
                st.session_state.who_we_are = DEFAULT_WHO_WE_ARE
                st.session_state.grade_rubric = DEFAULT_GRADE_RUBRIC
                st.rerun()

        st.session_state.workers = st.slider("Concurrent agents", 1, 10, st.session_state.get("workers", 6))
        only_ungraded = st.checkbox("Only grade ungraded rows", value=True)
        limit_on = st.checkbox("Limit to first N rows (test run)")
        row_limit = st.number_input("N", min_value=1, value=5, step=1) if limit_on else None

        st.divider()
        pending_count = len(build_tasks(only_ungraded, row_limit))
        st.caption(f"This run will grade **{pending_count}** compan{'y' if pending_count == 1 else 'ies'}.")
        disabled = pending_count == 0 or not get_api_key()
        if not get_api_key():
            st.warning("CURSOR_API_KEY is not set (env var or st.secrets).")
        if st.button("🚀 Start Grading", disabled=disabled, type="primary"):
            start_run(build_tasks(only_ungraded, row_limit))
            st.rerun()


# ---------------------------------------------------------------------------
# 3. Live run -- Mission Control
# ---------------------------------------------------------------------------

def drain_engine_events() -> bool:
    """Pull everything currently queued and fold it into session_state.
    Returns True once the engine has signalled it's fully finished."""
    engine: GradingEngine = st.session_state.engine
    finished = False
    while True:
        try:
            event = engine.events.get_nowait()
        except Empty:
            break
        etype = event["type"]

        if etype == "row_started":
            slot = event["slot"]
            st.session_state.worker_slots[slot] = {
                "row_index": event["row_index"], "company": event["company"] or event["website"],
                "website": event["website"], "lines": deque([{"kind": "system", "text": "Starting…"}], maxlen=200),
                "state": "running",
            }

        elif etype == "delta":
            slot = event["slot"]
            card = st.session_state.worker_slots.get(slot)
            if card and card["row_index"] == event["row_index"] and event["detail"]:
                text = event["detail"].strip().replace("\n", " ")
                kind = event["kind"]
                lines = card["lines"]
                # Same-kind consecutive chunks are one continuous thought streaming in --
                # merge them instead of showing each throttled flush as its own line,
                # which is what made reasoning read as disconnected fragments.
                if lines and lines[-1]["kind"] == kind and kind in MERGEABLE_DELTA_KINDS:
                    lines[-1]["text"] = f"{lines[-1]['text']} {text}".strip()
                else:
                    lines.append({"kind": kind, "text": text})

        elif etype == "row_finished":
            slot = event["slot"]
            row_index = event["row_index"]
            grade, reason = event["grade"], event["reason"]
            write_result_to_sheet(row_index, grade, reason)
            st.session_state.run_done += 1
            st.session_state.run_counts[grade] = st.session_state.run_counts.get(grade, 0) + 1
            company = event["company"] or event["website"]
            st.session_state.activity_feed.appendleft({
                "row_index": row_index, "company": company, "grade": grade, "reason": reason,
            })
            # Keep the card in place (marked complete) instead of clearing it --
            # it only gets replaced once this slot actually picks up a new company,
            # so the finished transcript stays visible in its original spot.
            card = st.session_state.worker_slots.get(slot)
            if card and card["row_index"] == row_index:
                card["state"] = "complete"
                card["grade"] = grade
                card["reason"] = reason

        elif etype == "row_skipped":
            st.session_state.run_done += 1

        elif etype == "engine_error":
            st.session_state.engine_error = event["error"]

        elif etype == "batch_finished":
            finished = True

    return finished


def render_transcript(lines: list[dict]) -> None:
    for line in lines:
        if line["kind"] == "system":
            st.markdown(f"**{line['text']}**")
        else:
            prefix = PHASE_PREFIX.get(line["kind"], "Working")
            st.markdown(f"**{prefix}:** {line['text']}")


@st.fragment(run_every="1s")
def render_mission_control() -> None:
    finished = drain_engine_events()

    total = st.session_state.run_total
    done = st.session_state.run_done
    elapsed = time.time() - st.session_state.run_start_time
    rate = (done / elapsed * 60) if elapsed > 0 else 0.0

    top_left, top_right = st.columns([1, 2])
    with top_left:
        st.metric("Graded", f"{done}/{total}")
        st.progress(done / total if total else 1.0)
    with top_right:
        counts = st.session_state.run_counts
        b1, b2, b3 = st.columns(3)
        with b1:
            st.badge(f"GOOD {counts.get('GOOD', 0)}", color="green")
        with b2:
            st.badge(f"MAYBE {counts.get('MAYBE', 0)}", color="orange")
        with b3:
            st.badge(f"UNABLE {counts.get('UNABLE', 0)}", color="gray")
        st.caption(f"{elapsed / 60:.1f} min elapsed · ~{rate:.1f} companies/min · keep this tab open")

    if st.session_state.get("engine_error"):
        st.error(f"Engine error: {st.session_state.engine_error}")

    st.subheader("Cursor Agents")
    slots = st.session_state.worker_slots
    cols = st.columns(min(len(slots), 3) or 1)
    for i, slot_id in enumerate(sorted(slots)):
        card = slots[slot_id]
        with cols[i % len(cols)]:
            if card is None:
                st.caption(f"**Agent {slot_id + 1}:** waiting for next company…")
            else:
                # A real widget (key=) so the open/closed choice survives every
                # 1s fragment rerun -- st.status's own chevron click doesn't,
                # since we recreate it fresh each tick. The card itself stays in
                # place (state flips to "complete") once its company finishes,
                # instead of being cleared, so the transcript stays readable
                # until this slot actually picks up a new company.
                expand_key = f"expand_agent_{slot_id}"
                # Agent 1 defaults open (via the widget's own first-run value=,
                # not an external session_state pre-seed, which didn't reliably
                # survive repeated run_every fragment reruns) so it's obvious
                # from the start that these boxes are toggleable.
                default_open = slot_id == 0 and expand_key not in st.session_state
                expanded = st.checkbox(
                    f"**Agent {slot_id + 1}:** {card['company']}",
                    value=default_open, key=expand_key,
                )
                status_state = card.get("state", "running")
                with st.status(card["website"], state=status_state, expanded=expanded, type="compact"):
                    if status_state == "complete":
                        st.badge(card["grade"], color=GRADE_BADGE_COLOR.get(card["grade"], "gray"))
                        st.write(card["reason"])
                    log_box = st.container(height=260)
                    with log_box:
                        render_transcript(card["lines"])

    st.subheader("Recently graded")
    if st.session_state.activity_feed:
        for item in st.session_state.activity_feed:
            row = st.columns([1, 5])
            with row[0]:
                st.badge(item["grade"], color=GRADE_BADGE_COLOR.get(item["grade"], "gray"))
            with row[1]:
                st.write(f"**{item['company']}** — {item['reason']}")
    else:
        st.caption("Nothing graded yet.")

    if finished:
        st.session_state.run_finished = True

    if st.session_state.get("run_finished"):
        st.success("Grading run complete.")
        if st.button("View results →", type="primary"):
            st.session_state.stage = "review"
            st.rerun()
    elif st.button("⏹ Stop"):
        st.session_state.engine.stop()
        st.info("Stopping — in-flight companies will finish, no new ones will start.")


# ---------------------------------------------------------------------------
# 4. Review & export
# ---------------------------------------------------------------------------

def reset_session() -> None:
    for key in list(st.session_state.keys()):
        del st.session_state[key]


def render_review() -> None:
    top_l, top_r = st.columns([5, 1])
    with top_l:
        st.title("✅ Grading complete")
    with top_r:
        if st.button("↩ Start Over", use_container_width=True, type="primary"):
            reset_session()
            st.rerun()
    df = dataframe_from_ws()

    c1, c2 = st.columns(2)
    with c1:
        grade_filter = st.multiselect("Filter by grade", options=sorted(VALID_GRADES), default=[])
    with c2:
        search = st.text_input("Search company / website")

    view = df
    if grade_filter:
        view = view[view["WEBSITE_GRADE"].isin(grade_filter)]
    if search:
        mask = view.apply(lambda r: search.lower() in str(r.get("Company Name", "")).lower()
                           or search.lower() in str(r.get("Website Link", "")).lower(), axis=1)
        view = view[mask]

    edited = st.data_editor(
        view, use_container_width=True, height=450, hide_index=True,
        disabled=[c for c in df.columns if c not in OUTPUT_COLUMNS],
        key="review_editor",
    )

    if st.button("💾 Save edits back to sheet"):
        header = st.session_state.header
        # Match edited rows back to sheet rows by Website Link (unique per row in this sheet).
        by_website = {str(r["Website Link"]): r for _, r in edited.iterrows()}
        for row in st.session_state.ws.iter_rows(min_row=2):
            website = row[header["Website Link"] - 1].value
            if website is None or str(website) not in by_website:
                continue
            r = by_website[str(website)]
            row[header["WEBSITE_GRADE"] - 1].value = r["WEBSITE_GRADE"]
            row[header["Comment"] - 1].value = r["Comment"]
        st.success("Saved.")

    st.download_button(
        "⬇️ Download graded .xlsx", data=workbook_bytes(),
        file_name=f"graded_{st.session_state.get('uploaded_name', 'companies.xlsx')}",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    unable_indices = set()
    header = st.session_state.header
    for row in st.session_state.ws.iter_rows(min_row=2):
        if str(row[header["WEBSITE_GRADE"] - 1].value or "").strip().upper() == "UNABLE":
            unable_indices.add(row[0].row)
    if unable_indices and st.button(f"🔁 Re-run {len(unable_indices)} UNABLE rows"):
        start_run(build_tasks(only_ungraded=False, row_limit=None, only_row_indices=unable_indices))
        st.rerun()


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    if not check_password():
        return

    stage = st.session_state.get("stage", "upload")
    if "wb" not in st.session_state:
        render_upload()
    elif stage == "setup":
        render_setup()
    elif stage == "running":
        render_mission_control()
    elif stage == "review":
        render_review()


if __name__ == "__main__":
    main()
